"""crosscorrelation_tab.py - GUI tab for cluster x cluster cross-correlation.

Lives next to the Clustering tab. Two modes:
  - Full recording: run cross-correlation on each ROI pair using the
    entire dF/F trace.
  - Per event: re-run cross-correlation on each ROI pair, cropped to each
    detected event window.

Both modes use the batched cross-correlation in
``crosscorrelation.batch_xcorr_clusters``: one matmul per lag covers every
ROI pair in the cluster pair, so we never build the full +/- max_lag
distribution per pair. Outputs are CSV summaries with ``best_lag_sec``,
``max_corr`` (Pearson r at peak) and ``zero_lag_corr`` (Pearson r at 0).
"""

from __future__ import annotations

import queue
import sys
import threading
import tkinter as tk
import traceback
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Optional

import numpy as np
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg, NavigationToolbar2Tk,
)

sys.path.insert(0, str(Path(__file__).resolve().parent))
import crosscorrelation as xc  # noqa: E402
import utils  # noqa: E402


DEFAULT_PREFIX = "r0p7_filtered_"
DEFAULT_CLUSTER_FOLDER = "gui_recluster"
DEFAULT_MAX_LAG_S = 2.0
POLL_MS = 100


def _read_event_windows_from_summary(plane0: Path):
    """Pull (start_s, end_s) tuples from the EventWindows sheet written by
    ``summary_writer.write_events_sheets``. Returns [] if the sheet or
    workbook is missing.
    """
    try:
        import openpyxl  # noqa
    except Exception:
        return []
    candidates = sorted(plane0.glob("*summary*.xlsx"))
    for path in candidates:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        if "EventWindows" not in wb.sheetnames:
            wb.close()
            continue
        ws = wb["EventWindows"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            continue
        header = [str(c).strip().lower() if c is not None else ""
                  for c in rows[0]]
        try:
            i_s = header.index("start_s")
            i_e = header.index("end_s")
        except ValueError:
            continue
        out = []
        for r in rows[1:]:
            try:
                out.append((float(r[i_s]), float(r[i_e])))
            except (TypeError, ValueError):
                continue
        if out:
            return out
    return []


class CrossCorrelationTab(ttk.Frame):
    """Cluster x cluster cross-correlation, full recording + per-event."""

    def __init__(self, master, state=None) -> None:
        super().__init__(master, padding=10)
        self.state = state

        self._plane0: Optional[Path] = None
        self._prefix = DEFAULT_PREFIX
        self._cluster_folder = DEFAULT_CLUSTER_FOLDER
        self._fps: float = 15.07
        self._event_windows: list[tuple[float, float]] = []

        # Cache for the single-pair preview plot (avoid reopening memmap).
        self._dff_cache: Optional[np.memmap] = None
        self._dff_cache_key = None  # (plane0, prefix)

        self._q: queue.Queue = queue.Queue()
        self._worker: Optional[threading.Thread] = None

        self._build_ui()
        self.after(POLL_MS, self._drain_queue)

        if state is not None:
            try:
                state.subscribe_plane0(self._on_plane0_broadcast)
                state.subscribe_lowpass_ready(self._on_plane0_broadcast)
                if getattr(state, "lowpass_plane0", None) is not None:
                    self._on_plane0_broadcast(state.lowpass_plane0)
                elif getattr(state, "plane0", None) is not None:
                    self._on_plane0_broadcast(state.plane0)
            except Exception:
                pass

    # -- UI ----------------------------------------------------------------

    def _build_ui(self) -> None:
        head = ttk.LabelFrame(
            self, text="Cluster x cluster cross-correlation "
                       "(batched matmul, best lag + zero lag)", padding=8)
        head.pack(fill="x", pady=(0, 6))

        row1 = ttk.Frame(head); row1.pack(fill="x", pady=2)
        ttk.Label(row1, text="plane0:").pack(side="left")
        self.path_var = tk.StringVar(value="")
        ttk.Entry(row1, textvariable=self.path_var, width=70).pack(
            side="left", padx=(4, 4), fill="x", expand=True)
        ttk.Button(row1, text="Browse...",
                   command=self._on_browse).pack(side="left")

        row2 = ttk.Frame(head); row2.pack(fill="x", pady=2)
        ttk.Label(row2, text="prefix:").pack(side="left")
        self.prefix_var = tk.StringVar(value=DEFAULT_PREFIX)
        ttk.Entry(row2, textvariable=self.prefix_var, width=18).pack(
            side="left", padx=(4, 12))

        ttk.Label(row2, text="cluster folder:").pack(side="left")
        self.cfolder_var = tk.StringVar(value=DEFAULT_CLUSTER_FOLDER)
        ttk.Entry(row2, textvariable=self.cfolder_var, width=18).pack(
            side="left", padx=(4, 12))

        ttk.Label(row2, text="fps:").pack(side="left")
        self.fps_var = tk.StringVar(value="15.07")
        ttk.Entry(row2, textvariable=self.fps_var, width=8).pack(
            side="left", padx=(4, 0))

        # Algorithm parameters
        row3 = ttk.LabelFrame(head, text="Search parameters", padding=6)
        row3.pack(fill="x", pady=(6, 0))

        ttk.Label(row3, text="max_lag (s):").pack(side="left")
        self.maxlag_var = tk.StringVar(value=str(DEFAULT_MAX_LAG_S))
        ttk.Entry(row3, textvariable=self.maxlag_var, width=6).pack(
            side="left", padx=(4, 12))

        self.zero_lag_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(row3, text="also output zero-lag corr",
                        variable=self.zero_lag_var).pack(
            side="left", padx=(4, 12))

        self.gpu_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(row3, text="use GPU if available",
                        variable=self.gpu_var).pack(side="left")

        # Run buttons
        run_frame = ttk.LabelFrame(self, text="Run", padding=8)
        run_frame.pack(fill="x", pady=(0, 6))

        self.full_btn = ttk.Button(
            run_frame, text="Run full-recording cross-correlation",
            command=self._on_run_full, state="disabled")
        self.full_btn.pack(side="left")

        self.per_event_btn = ttk.Button(
            run_frame, text="Run per-event cross-correlation",
            command=self._on_run_per_event, state="disabled")
        self.per_event_btn.pack(side="left", padx=(8, 0))

        self.refresh_btn = ttk.Button(
            run_frame, text="Reload event windows",
            command=self._on_refresh_events)
        self.refresh_btn.pack(side="left", padx=(8, 0))

        self.event_count_var = tk.StringVar(value="events: -")
        ttk.Label(run_frame, textvariable=self.event_count_var).pack(
            side="left", padx=(12, 0))

        # Progress + status
        prog = ttk.Frame(self); prog.pack(fill="x", pady=(0, 6))
        self.progress = ttk.Progressbar(prog, mode="determinate", length=320)
        self.progress.pack(side="left")
        self.status_var = tk.StringVar(value="Pick a plane0 folder.")
        ttk.Label(prog, textvariable=self.status_var,
                  font=("", 9, "italic")).pack(
            side="left", padx=(8, 0), fill="x", expand=True)

        # Bottom: log on the left, single-pair preview on the right.
        bottom = ttk.PanedWindow(self, orient="horizontal")
        bottom.pack(fill="both", expand=True)

        # -- Log pane
        log_frame = ttk.LabelFrame(bottom, text="Log", padding=4)
        bottom.add(log_frame, weight=1)
        self.log_text = tk.Text(log_frame, height=12, wrap="word",
                                font=("Consolas", 9))
        self.log_text.pack(fill="both", expand=True, side="left")
        sb = ttk.Scrollbar(log_frame, orient="vertical",
                           command=self.log_text.yview)
        sb.pack(fill="y", side="right")
        self.log_text.config(yscrollcommand=sb.set)

        # -- Single-pair preview pane
        sp_frame = ttk.LabelFrame(
            bottom, text="Single-pair cross-correlation curve", padding=6)
        bottom.add(sp_frame, weight=2)

        ctrl = ttk.Frame(sp_frame); ctrl.pack(fill="x", pady=(0, 4))
        ttk.Label(ctrl, text="ROI A:").pack(side="left")
        self.sp_roiA_var = tk.StringVar(value="0")
        ttk.Entry(ctrl, textvariable=self.sp_roiA_var, width=7).pack(
            side="left", padx=(2, 8))
        ttk.Label(ctrl, text="ROI B:").pack(side="left")
        self.sp_roiB_var = tk.StringVar(value="1")
        ttk.Entry(ctrl, textvariable=self.sp_roiB_var, width=7).pack(
            side="left", padx=(2, 8))
        ttk.Label(ctrl, text="lag window (s):").pack(side="left")
        self.sp_lag_var = tk.StringVar(value=str(DEFAULT_MAX_LAG_S))
        ttk.Entry(ctrl, textvariable=self.sp_lag_var, width=6).pack(
            side="left", padx=(2, 8))
        self.sp_event_var = tk.StringVar(value="full recording")
        ttk.Label(ctrl, text="event:").pack(side="left")
        self.sp_event_combo = ttk.Combobox(
            ctrl, textvariable=self.sp_event_var, state="readonly",
            width=22, values=["full recording"])
        self.sp_event_combo.pack(side="left", padx=(2, 8))
        self.sp_plot_btn = ttk.Button(
            ctrl, text="Plot", command=self._on_plot_single_pair,
            state="disabled")
        self.sp_plot_btn.pack(side="left")

        self.sp_fig = plt.Figure(figsize=(5.5, 3.0), tight_layout=True)
        self.sp_ax = self.sp_fig.add_subplot(111)
        self._sp_placeholder("Pick two ROIs and click Plot.")
        self.sp_canvas = FigureCanvasTkAgg(self.sp_fig, master=sp_frame)
        tb_frame = ttk.Frame(sp_frame); tb_frame.pack(fill="x")
        self.sp_toolbar = NavigationToolbar2Tk(
            self.sp_canvas, tb_frame, pack_toolbar=False)
        self.sp_toolbar.update()
        self.sp_toolbar.pack(side="left", fill="x")
        self.sp_canvas.get_tk_widget().pack(fill="both", expand=True)

    # -- Single-pair preview helpers ---------------------------------------

    def _sp_placeholder(self, text: str) -> None:
        self.sp_ax.clear()
        self.sp_ax.set_axis_off()
        self.sp_ax.text(0.5, 0.5, text, ha="center", va="center",
                        transform=self.sp_ax.transAxes)

    def _get_cached_dff(self):
        """Return (dff_memmap, T, N_kept), reopening only when path/prefix
        changes."""
        if self._plane0 is None:
            raise RuntimeError("No plane0 selected.")
        prefix = self.prefix_var.get().strip() or DEFAULT_PREFIX
        key = (str(self._plane0), prefix)
        if self._dff_cache is None or self._dff_cache_key != key:
            dff, T, N = xc._open_dff_memmap(self._plane0, prefix)
            self._dff_cache = (dff, T, N)
            self._dff_cache_key = key
        return self._dff_cache

    def _on_plot_single_pair(self) -> None:
        try:
            dff, T, N = self._get_cached_dff()
        except Exception as e:
            messagebox.showerror("dF/F unavailable", str(e))
            return
        # Parse ROI indices
        try:
            iA = int(self.sp_roiA_var.get())
            iB = int(self.sp_roiB_var.get())
        except ValueError:
            messagebox.showerror(
                "Bad ROI index", "ROI A and ROI B must be integers.")
            return
        if not (0 <= iA < N and 0 <= iB < N):
            messagebox.showerror(
                "Out of range", f"ROI indices must be in [0, {N - 1}].")
            return
        try:
            max_lag = float(self.sp_lag_var.get())
        except ValueError:
            max_lag = DEFAULT_MAX_LAG_S
        try:
            fps = float(self.fps_var.get())
        except ValueError:
            fps = self._fps

        # Resolve event crop (or full recording).
        ev_label = self.sp_event_var.get()
        f0, f1 = 0, T
        ev_title = "full recording"
        if ev_label != "full recording":
            try:
                ev_idx = int(ev_label.split()[1])
                s_sec, e_sec = self._event_windows[ev_idx]
                f0 = max(0, int(round(s_sec * fps)))
                f1 = min(T, int(round(e_sec * fps)))
                ev_title = (f"event {ev_idx:04d}  [{s_sec:.2f}-{e_sec:.2f}s, "
                            f"{f1 - f0} frames]")
            except (ValueError, IndexError):
                ev_title = "full recording"

        if f1 - f0 < 4:
            messagebox.showerror(
                "Window too short",
                f"Selected window has only {f1 - f0} frames.")
            return
        # Cap max_lag to window length.
        eff_max_lag = min(max_lag, max(0.0, (f1 - f0 - 1) / float(fps)))
        if eff_max_lag <= 0:
            messagebox.showerror(
                "Lag too large",
                "Lag window is wider than the available data window.")
            return

        sigA = np.asarray(dff[f0:f1, iA], dtype=np.float32)
        sigB = np.asarray(dff[f0:f1, iB], dtype=np.float32)
        try:
            lags_sec, r = xc.single_pair_xcorr_curve(
                sigA, sigB, fps, max_lag_seconds=eff_max_lag)
        except Exception as e:
            messagebox.showerror("xcorr failed", str(e))
            return

        peak_idx = int(np.argmax(r))
        peak_lag = float(lags_sec[peak_idx])
        peak_r = float(r[peak_idx])
        zero_idx = int(np.argmin(np.abs(lags_sec)))
        zero_r = float(r[zero_idx])

        ax = self.sp_ax
        ax.clear()
        ax.set_axis_on()
        ax.plot(lags_sec, r, lw=1.2, color="tab:blue")
        ax.axhline(0, color="black", lw=0.5, alpha=0.4)
        ax.axvline(0, color="black", lw=0.5, alpha=0.4)
        ax.axvline(peak_lag, color="tab:red", lw=1.0, ls="--",
                   label=f"peak: lag={peak_lag:.3f}s, r={peak_r:.3f}")
        ax.scatter([0.0], [zero_r], color="tab:green", zorder=5,
                   label=f"lag=0:  r={zero_r:.3f}")
        ax.set_xlabel("lag (s)  [+ = ROI A leads]")
        ax.set_ylabel("Pearson r (biased)")
        ax.set_title(f"ROI {iA}  vs  ROI {iB}  -  {ev_title}",
                     fontsize=10)
        ax.legend(fontsize=8, loc="best", frameon=True)
        ax.grid(True, alpha=0.3)
        self.sp_canvas.draw_idle()
        try:
            self.sp_toolbar.update()
        except Exception:
            pass
        self._log(f"[single-pair] ROI {iA} x ROI {iB}  ({ev_title})  "
                  f"peak_lag={peak_lag:.3f}s  peak_r={peak_r:.3f}  "
                  f"zero_r={zero_r:.3f}")

    # -- State plumbing ----------------------------------------------------

    def _on_plane0_broadcast(self, plane0) -> None:
        if plane0 is None:
            return
        self.path_var.set(str(plane0))
        self._set_plane0(Path(plane0))

    def _on_browse(self) -> None:
        path = filedialog.askdirectory(
            title="Select Suite2p plane0 folder",
            initialdir=self.path_var.get() or str(Path.home()))
        if not path:
            return
        self.path_var.set(path)
        self._set_plane0(Path(path))

    def _set_plane0(self, plane0: Path) -> None:
        self._plane0 = plane0
        # Reset the dF/F memmap cache so the next preview reopens.
        self._dff_cache = None
        self._dff_cache_key = None
        try:
            fps = float(utils.get_fps_from_notes(str(plane0),
                                                 default_fps=15.07))
            self._fps = fps
            self.fps_var.set(f"{fps:.3f}")
        except Exception:
            pass
        self._on_refresh_events()
        ok = self._inputs_ready()
        self.full_btn.config(state="normal" if ok else "disabled")
        self.per_event_btn.config(
            state="normal" if (ok and self._event_windows) else "disabled")
        # Single-pair preview only needs the dF/F memmap, not the cluster files.
        prefix = self.prefix_var.get().strip() or DEFAULT_PREFIX
        sp_ok = (plane0 is not None
                 and (plane0 / f"{prefix}dff.memmap.float32").exists())
        self.sp_plot_btn.config(state="normal" if sp_ok else "disabled")
        if ok:
            self.status_var.set(f"Ready. ({plane0})")
        else:
            self.status_var.set("Cluster files or dF/F memmap missing.")

    def _inputs_ready(self) -> bool:
        if self._plane0 is None:
            return False
        prefix = self.prefix_var.get().strip() or DEFAULT_PREFIX
        cfolder = self.cfolder_var.get().strip() or DEFAULT_CLUSTER_FOLDER
        cluster_dir = (self._plane0 / f"{prefix}cluster_results" / cfolder)
        if not (self._plane0 / f"{prefix}dff.memmap.float32").exists():
            return False
        if not cluster_dir.exists():
            return False
        roi_files = [f for f in cluster_dir.glob("*_rois.npy")
                     if "manual_combined" not in f.stem.lower()]
        return len(roi_files) >= 1

    def _on_refresh_events(self) -> None:
        if self._plane0 is None:
            self._event_windows = []
            self.event_count_var.set("events: -")
            self.sp_event_combo.config(values=["full recording"])
            self.sp_event_var.set("full recording")
            return
        evts = _read_event_windows_from_summary(self._plane0)
        self._event_windows = evts
        self.event_count_var.set(f"events: {len(evts)}")
        self.per_event_btn.config(
            state="normal" if (self._inputs_ready() and evts) else "disabled")
        # Refresh the single-pair event picker.
        labels = ["full recording"] + [
            f"event {i:04d}  [{s:.2f}-{e:.2f}s]"
            for i, (s, e) in enumerate(evts)
        ]
        self.sp_event_combo.config(values=labels)
        if self.sp_event_var.get() not in labels:
            self.sp_event_var.set("full recording")

    # -- Logging -----------------------------------------------------------

    def _log(self, msg: str) -> None:
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")

    def _set_progress(self, done: int, total: int) -> None:
        if total <= 0:
            self.progress.config(value=0, maximum=1)
            return
        self.progress.config(value=done, maximum=total)

    # -- Run: full recording -----------------------------------------------

    def _gather_params(self):
        prefix = self.prefix_var.get().strip() or DEFAULT_PREFIX
        cfolder = self.cfolder_var.get().strip() or DEFAULT_CLUSTER_FOLDER
        try:
            fps = float(self.fps_var.get())
        except ValueError:
            fps = self._fps
        try:
            max_lag = float(self.maxlag_var.get())
        except ValueError:
            max_lag = DEFAULT_MAX_LAG_S
        return {
            "prefix": prefix, "cluster_folder": cfolder, "fps": fps,
            "max_lag_seconds": max_lag,
            "zero_lag": bool(self.zero_lag_var.get()),
            "use_gpu": bool(self.gpu_var.get()),
        }

    def _disable_run(self) -> None:
        self.full_btn.config(state="disabled")
        self.per_event_btn.config(state="disabled")

    def _enable_run(self) -> None:
        ok = self._inputs_ready()
        self.full_btn.config(state="normal" if ok else "disabled")
        self.per_event_btn.config(
            state="normal" if (ok and self._event_windows) else "disabled")

    def _on_run_full(self) -> None:
        if self._plane0 is None or not self._inputs_ready():
            messagebox.showerror("Not ready", "Inputs incomplete.")
            return
        if self._worker is not None and self._worker.is_alive():
            messagebox.showinfo("Busy", "A run is already in progress.")
            return
        params = self._gather_params()
        plane0 = self._plane0
        self._disable_run()
        self.progress.config(value=0, maximum=1)
        self.status_var.set("Running full-recording cross-correlation...")
        self._log(f"\n[full] plane0={plane0}")
        self._log(f"[full] params={params}")

        def progress_cb(done, total, label):
            self._q.put(("progress", (done, total, label)))

        def worker():
            try:
                out = xc.run_cluster_xcorr_full_fast(
                    plane0, progress_cb=progress_cb, **params,
                )
                self._q.put(("done_full", str(out)))
            except Exception as e:
                self._q.put(("error", f"{e}\n{traceback.format_exc()}"))

        self._worker = threading.Thread(target=worker, daemon=True)
        self._worker.start()

    def _on_run_per_event(self) -> None:
        if self._plane0 is None or not self._inputs_ready():
            messagebox.showerror("Not ready", "Inputs incomplete.")
            return
        if not self._event_windows:
            messagebox.showinfo(
                "No events",
                "No event windows found in the recording summary "
                "(*summary*.xlsx). Run event detection first.")
            return
        if self._worker is not None and self._worker.is_alive():
            messagebox.showinfo("Busy", "A run is already in progress.")
            return
        params = self._gather_params()
        plane0 = self._plane0
        evts = list(self._event_windows)
        self._disable_run()
        self.progress.config(value=0, maximum=len(evts))
        self.status_var.set(
            f"Running per-event cross-correlation ({len(evts)} events)...")
        self._log(f"\n[per-event] plane0={plane0}")
        self._log(f"[per-event] params={params}  events={len(evts)}")

        def progress_cb(done, total, label):
            self._q.put(("progress", (done, total, label)))

        def worker():
            try:
                out = xc.run_cluster_xcorr_per_event_fast(
                    plane0, event_windows=evts,
                    progress_cb=progress_cb, **params,
                )
                self._q.put(("done_per_event", str(out)))
            except Exception as e:
                self._q.put(("error", f"{e}\n{traceback.format_exc()}"))

        self._worker = threading.Thread(target=worker, daemon=True)
        self._worker.start()

    # -- Queue drain -------------------------------------------------------

    def _drain_queue(self) -> None:
        try:
            while True:
                kind, payload = self._q.get_nowait()
                if kind == "progress":
                    done, total, label = payload
                    self._set_progress(done, total)
                    self.status_var.set(f"{label} ({done}/{total})")
                elif kind == "done_full":
                    self._set_progress(1, 1)
                    self.status_var.set(f"Done. Output: {payload}")
                    self._log(f"[full] done -> {payload}")
                    self._enable_run()
                elif kind == "done_per_event":
                    n = int(self.progress.cget("maximum"))
                    self._set_progress(n, n)
                    self.status_var.set(f"Done. Output: {payload}")
                    self._log(f"[per-event] done -> {payload}")
                    self._enable_run()
                elif kind == "error":
                    self.status_var.set("Cross-correlation failed.")
                    self._log("[error] " + payload)
                    self._enable_run()
                    messagebox.showerror(
                        "Cross-correlation failed",
                        payload.split("\n", 1)[0])
        except queue.Empty:
            pass
        self.after(POLL_MS, self._drain_queue)


def main() -> None:
    root = tk.Tk()
    root.title("CalLIOPE - Cross-correlation")
    root.geometry("1100x720")
    tab = CrossCorrelationTab(root, state=None)
    tab.pack(fill="both", expand=True)
    root.mainloop()


if __name__ == "__main__":
    main()
